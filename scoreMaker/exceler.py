import openpyxl

def saveExcel(title, ScoreListR, ScoreListL):
    ico = 0
    jco = 0
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'sheet_1'
    wb.save(title+'.xlsx')

    wb = openpyxl.load_workbook(title+'.xlsx')
    sheet = wb['sheet_1']
    
    for i in ScoreListR:
        for j in i:
            #row ,colm 1から
            sheet.cell(row=jco+1, column=ico+1, value=ScoreListR[ico][jco])
            jco+=1
        jco = 0
        ico+=1

    ico = 0
    jco = 0
    for i in ScoreListL:
        for j in i:
            #row ,colm 1から
            sheet.cell(row=6+jco+1, column=ico+1, value=ScoreListL[ico][jco])
            jco+=1
        jco = 0
        ico+=1
    
    wb.save(title+'.xlsx')

if __name__ == "__main__":
    ScoreListR = [[1,2,1,2,1],[1,2,4,1,2],[1,2,2,4,3],[1,2,8,2,3]]
    ScoreListL = [[1,2,1,2,1],[1,2,4,1,2],[1,2,2,4,3],[1,2,8,2,0]]
    title = "test"
    saveExcel(title, ScoreListR,ScoreListL)
